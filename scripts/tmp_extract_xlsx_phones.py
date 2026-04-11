from __future__ import annotations

import csv
import re
import sys
from pathlib import Path


def _stringy(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # Excel sometimes stores phone numbers as floats/ints; keep raw string but drop trailing ".0"
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: python scripts/tmp_extract_xlsx_phones.py <path-to-xlsx> [out-csv]")
        return 2

    xlsx_path = Path(sys.argv[1]).expanduser()
    out_csv = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("data/_tmp_lead_sheet_extract.csv")
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        print("details:", e)
        return 1

    if not xlsx_path.exists():
        print(f"ERROR: not found: {xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # Find first non-empty row as header.
    header_row_idx = None
    for i, r in enumerate(rows[:100]):
        if any(_stringy(c) for c in r):
            header_row_idx = i
            break
    if header_row_idx is None:
        print("ERROR: workbook appears empty")
        return 1

    raw_headers = [_stringy(c) for c in rows[header_row_idx]]
    headers = [h.lower() for h in raw_headers]

    def find_col(names: list[str]) -> int:
        for name in names:
            for j, h in enumerate(headers):
                if name in h:
                    return j
        return -1

    phone_col = find_col(["phone", "mobile", "telephone", "tel", "contact number", "phone number", "number"])
    name_col = find_col(["name", "full name", "contact", "contact name", "lead name", "decision maker", "decisionmaker"])
    service_col = find_col(["service", "interest", "product", "inquiry", "service type"])
    source_col = find_col(["source", "lead source", "campaign", "utm_source", "origin"])

    # Fallback: pick the column with the most phone-like values.
    if phone_col == -1:
        phone_like = re.compile(r"\+?\d[\d\s\-().]{6,}")
        best = (-1, -1)
        for j in range(len(raw_headers)):
            cnt = 0
            for r in rows[header_row_idx + 1 : header_row_idx + 401]:
                if j >= len(r):
                    continue
                if phone_like.search(_stringy(r[j])):
                    cnt += 1
            if cnt > best[0]:
                best = (cnt, j)
        phone_col = best[1] if best[0] > 0 else -1

    print("sheets:", wb.sheetnames)
    print("header_row:", header_row_idx + 1)
    print("detected_columns:", {
        "name": (raw_headers[name_col] if name_col != -1 else None),
        "phone": (raw_headers[phone_col] if phone_col != -1 else None),
        "service": (raw_headers[service_col] if service_col != -1 else None),
        "source": (raw_headers[source_col] if source_col != -1 else None),
    })

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["row", "name", "phone", "service", "source"])
        for excel_row_num, r in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
            if not any(_stringy(c) for c in r):
                continue
            name = _stringy(r[name_col]) if name_col != -1 and name_col < len(r) else ""
            phone = _stringy(r[phone_col]) if phone_col != -1 and phone_col < len(r) else ""
            service = _stringy(r[service_col]) if service_col != -1 and service_col < len(r) else ""
            source = _stringy(r[source_col]) if source_col != -1 and source_col < len(r) else ""
            w.writerow([excel_row_num, name, phone, service, source])

    print("wrote:", out_csv)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

